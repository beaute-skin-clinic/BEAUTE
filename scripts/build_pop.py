#!/usr/bin/env python3
"""
build_pop.py — エクセルから施術POP HTMLを生成

Usage:
  python3 build_pop.py <input.xlsx> [output.html]
"""
from __future__ import annotations
import json, sys, os
import openpyxl

# ─── 設定 ───────────────────────────────────────────
SKIP_VALUES = {'設定なし', '対象外', '', None}
PRICE_COLS = {
    'old_price':  6,   # 旧通常価格
    'old_first':  7,   # 旧初回価格
    'price':      8,   # 通常価格（新）
    'first':      9,   # 初回価格
    'repeat':    10,   # リピート
    'campaign':  11,   # CP
    'moni_full': 12,   # モニ全顔
    'moni_mask': 13,   # モニ目隠
    'moni_eye':  14,   # モニ目元
    'ticket_3':  15,   # 3回券
    'ticket_5':  16,   # 5回券
    'ticket_8':  17,   # 8回券
    'bp10':      18,   # BP10万
    'bp30':      19,   # BP30万
    'bp50':      20,   # BP50万
    'ep5':       21,   # EP5万
    'ep10':      22,   # EP10万
    'ep20':      23,   # EP20万
}

def parse_price(v):
    """値を数値に変換。設定なし/対象外/None → None"""
    if v is None:
        return None
    s = str(v).strip()
    if s in SKIP_VALUES:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None

def read_excel(path):
    """エクセルから全施術データを読み込み → 施術名でグループ化"""
    wb = openpyxl.load_workbook(path, data_only=True)

    # まずフラットに全行を読む
    from collections import OrderedDict
    groups = OrderedDict()  # key = (sheet, name)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 3:
            continue

        for r in range(3, ws.max_row + 1):
            name = ws.cell(r, 3).value
            if not name:
                continue

            price = parse_price(ws.cell(r, 8).value)
            if price is None:
                continue

            name_str = str(name).strip()
            sheet_str = sheet_name.strip()
            key = (sheet_str, name_str)

            prices = {}
            for pkey, col in PRICE_COLS.items():
                v = parse_price(ws.cell(r, col).value)
                if v is not None:
                    prices[pkey] = v

            variant = {
                'area': str(ws.cell(r, 4).value or '').strip(),
                'unit': str(ws.cell(r, 5).value or '').strip(),
                'prices': prices,
            }

            if key not in groups:
                groups[key] = {
                    'sheet': sheet_str,
                    'kubun': str(ws.cell(r, 1).value or '').strip(),
                    'category': str(ws.cell(r, 2).value or '').strip(),
                    'name': name_str,
                    'variants': [],
                }
            groups[key]['variants'].append(variant)

    # IDを振る
    treatments = []
    for i, g in enumerate(groups.values()):
        g['id'] = i
        treatments.append(g)

    return treatments

# ─── HTML テンプレート ────────────────────────────────
HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>施術POP ジェネレーター - BEAUTE SKIN CLINIC</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@300;400;500;600;700;900&display=swap" rel="stylesheet">
<style>
:root {
  --gold: #B8935A;
  --gold-light: #D4B87C;
  --bronze: #9E7B5B;
  --espresso: #5C3D2E;
  --deep-brown: #3A2518;
  --cream: #FAF6F1;
  --sand: #E8DDD3;
  --warm-white: #FFFCF8;
  --sage: #7A9E76;
  --sage-light: #E8F0E6;
  --blue: #2563EB;
  --blue-light: #EFF6FF;
  --red: #DC2626;
  --red-light: #FEF2F2;
  --cyan: #0891B2;
  --cyan-light: #ECFEFF;
  --purple: #7C3AED;
  --purple-light: #F5F3FF;
}

* { margin: 0; padding: 0; box-sizing: border-box; }

body {
  font-family: 'Noto Sans JP', sans-serif;
  background: #f5f5f5;
  color: var(--deep-brown);
}

/* ─── コントロールパネル ─── */
.no-print {
  /* 印刷時に完全に消すためのラッパー */
}
.control-panel {
  background: var(--espresso);
  color: #fff;
  padding: 16px 24px;
  position: sticky;
  top: 0;
  z-index: 100;
  box-shadow: 0 2px 12px rgba(0,0,0,0.2);
}
.control-row {
  display: flex;
  align-items: center;
  gap: 16px;
  flex-wrap: wrap;
}
.control-row h1 {
  font-size: 15px;
  font-weight: 600;
  margin-right: auto;
  letter-spacing: 0.5px;
}
.control-row h1 span { color: var(--gold-light); }

.search-box {
  padding: 8px 14px;
  border: 1px solid rgba(255,255,255,0.2);
  border-radius: 8px;
  background: rgba(255,255,255,0.1);
  color: #fff;
  font-size: 14px;
  width: 300px;
  font-family: inherit;
}
.search-box::placeholder { color: rgba(255,255,255,0.5); }
.search-box:focus { outline: none; border-color: var(--gold); background: rgba(255,255,255,0.15); }

.cat-filter {
  display: flex;
  gap: 6px;
  flex-wrap: wrap;
  margin-top: 10px;
}
.cat-btn {
  padding: 4px 12px;
  border: 1px solid rgba(255,255,255,0.25);
  border-radius: 16px;
  background: transparent;
  color: rgba(255,255,255,0.7);
  font-size: 11px;
  cursor: pointer;
  font-family: inherit;
  transition: all 0.15s;
}
.cat-btn:hover { background: rgba(255,255,255,0.1); color: #fff; }
.cat-btn.active { background: var(--gold); border-color: var(--gold); color: #000; font-weight: 600; }

.action-row {
  display: flex;
  gap: 10px;
  align-items: center;
  margin-top: 10px;
}
.layout-btn {
  padding: 6px 14px;
  border: 1px solid rgba(255,255,255,0.25);
  border-radius: 6px;
  background: transparent;
  color: #fff;
  font-size: 12px;
  cursor: pointer;
  font-family: inherit;
  transition: all 0.15s;
}
.layout-btn:hover { background: rgba(255,255,255,0.1); }
.layout-btn.active { background: var(--gold); border-color: var(--gold); color: #000; font-weight: 600; }

.btn-print {
  padding: 8px 20px;
  background: var(--gold);
  border: none;
  border-radius: 6px;
  color: #000;
  font-size: 13px;
  font-weight: 700;
  cursor: pointer;
  font-family: inherit;
  margin-left: auto;
}
.btn-print:hover { background: var(--gold-light); }

.selected-count {
  font-size: 12px;
  color: var(--gold-light);
}
.btn-clear {
  padding: 4px 12px;
  border: 1px solid rgba(255,255,255,0.25);
  border-radius: 6px;
  background: transparent;
  color: rgba(255,255,255,0.7);
  font-size: 11px;
  cursor: pointer;
  font-family: inherit;
}
.btn-clear:hover { background: rgba(255,255,255,0.1); color: #fff; }

/* ─── 施術リスト（選択用） ─── */
.list-panel {
  padding: 16px 24px;
  max-height: 260px;
  overflow-y: auto;
  background: #fff;
  border-bottom: 1px solid #e5e5e5;
}
.list-grid {
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
  gap: 6px;
}
.list-item {
  display: flex;
  align-items: center;
  gap: 8px;
  padding: 6px 10px;
  border-radius: 6px;
  cursor: pointer;
  font-size: 12px;
  transition: background 0.1s;
  border: 1px solid transparent;
}
.list-item:hover { background: var(--cream); }
.list-item.selected { background: #FEF3C7; border-color: #FCD34D; }
.list-item .li-check {
  width: 18px; height: 18px;
  border: 2px solid #ccc;
  border-radius: 4px;
  flex-shrink: 0;
  display: flex; align-items: center; justify-content: center;
  font-size: 12px;
  transition: all 0.15s;
}
.list-item.selected .li-check { background: var(--gold); border-color: var(--gold); color: #fff; }
.list-item .li-name { font-weight: 600; }
.list-item .li-area { color: #888; margin-left: 4px; }
.list-item .li-price { margin-left: auto; font-weight: 700; color: var(--espresso); white-space: nowrap; }

/* ─── カードプレビュー（画面用） ─── */
.preview-area {
  padding: 24px;
}
.preview-hint {
  text-align: center;
  color: #999;
  font-size: 14px;
  padding: 60px 0;
}

/* カードグリッド — 可変高さで自然に流れる */
.card-grid {
  display: grid;
  gap: 16px;
  align-items: start;
}
.card-grid.layout-1 { grid-template-columns: 1fr; max-width: 600px; margin: 0 auto; }
.card-grid.layout-2 { grid-template-columns: repeat(2, 1fr); }
.card-grid.layout-4 { grid-template-columns: repeat(2, 1fr); }
.card-grid.layout-6 { grid-template-columns: repeat(3, 1fr); gap: 12px; }
.card-grid.layout-9 { grid-template-columns: repeat(3, 1fr); gap: 10px; }

/* ─── カード本体 ─── */
.pop-card {
  border: 2px solid var(--gold);
  border-radius: 12px;
  background: #fff;
  padding: 16px 16px 12px;
  display: flex;
  flex-direction: column;
}
.layout-4 .pop-card { padding: 10px 10px 8px; border-radius: 8px; }
.layout-6 .pop-card { padding: 7px 8px 6px; border-radius: 6px; border-width: 1.5px; }
.layout-9 .pop-card { padding: 4px 5px 3px; border-radius: 5px; border-width: 1.5px; }

.card-header {
  text-align: center;
  margin-bottom: 4px;
}
.card-clinic {
  font-size: 8px;
  letter-spacing: 2px;
  color: var(--bronze);
  font-weight: 500;
}
.layout-4 .card-clinic { font-size: 7px; }
.layout-6 .card-clinic { font-size: 6px; letter-spacing: 1.5px; }
.layout-9 .card-clinic { font-size: 5px; letter-spacing: 1px; }

.card-category {
  display: inline-block;
  padding: 1px 10px;
  background: var(--sage-light);
  color: var(--sage);
  border-radius: 20px;
  font-size: 8px;
  font-weight: 600;
  margin-top: 3px;
}
.layout-4 .card-category { font-size: 7px; padding: 1px 8px; margin-top: 2px; }
.layout-6 .card-category { font-size: 6px; padding: 1px 6px; margin-top: 1px; }
.layout-9 .card-category { display: none; }

.card-name {
  font-size: 20px;
  font-weight: 900;
  text-align: center;
  margin: 4px 0 2px;
  color: var(--deep-brown);
  line-height: 1.2;
}
.layout-1 .card-name { font-size: 30px; margin: 10px 0 4px; }
.layout-4 .card-name { font-size: 14px; margin: 2px 0 1px; }
.layout-6 .card-name { font-size: 12px; margin: 2px 0 0; }
.layout-9 .card-name { font-size: 9px; margin: 1px 0 0; }

.card-area {
  text-align: center;
  font-size: 13px;
  color: #888;
  margin-bottom: 2px;
}
.layout-4 .card-area { font-size: 10px; }
.layout-6 .card-area { font-size: 9px; }
.layout-9 .card-area { font-size: 7px; }

.card-unit {
  text-align: center;
  font-size: 11px;
  color: #aaa;
  margin-bottom: 10px;
}
.layout-4 .card-unit { font-size: 9px; margin-bottom: 6px; }
.layout-6 .card-unit { font-size: 8px; margin-bottom: 4px; }
.layout-9 .card-unit { font-size: 6px; margin-bottom: 2px; }

/* メイン価格 */
.main-price {
  text-align: center;
  margin-bottom: 10px;
  padding-bottom: 10px;
  border-bottom: 1px dashed var(--sand);
}
.layout-4 .main-price { margin-bottom: 6px; padding-bottom: 6px; }
.layout-6 .main-price { margin-bottom: 4px; padding-bottom: 4px; }
.layout-9 .main-price { margin-bottom: 2px; padding-bottom: 2px; }
.main-price-label { font-size: 11px; color: #888; margin-bottom: 2px; }
.layout-4 .main-price-label { font-size: 9px; }
.layout-6 .main-price-label { font-size: 8px; }
.layout-9 .main-price-label { font-size: 6px; margin-bottom: 0; }
.main-price-value {
  font-size: 34px;
  font-weight: 900;
  color: var(--deep-brown);
  letter-spacing: -1px;
}
.layout-1 .main-price-value { font-size: 48px; }
.layout-4 .main-price-value { font-size: 22px; }
.layout-6 .main-price-value { font-size: 18px; }
.layout-9 .main-price-value { font-size: 14px; letter-spacing: 0; }
.main-price-yen { font-size: 18px; font-weight: 700; }
.layout-4 .main-price-yen { font-size: 12px; }
.layout-6 .main-price-yen { font-size: 10px; }
.layout-9 .main-price-yen { font-size: 8px; }
.main-price-tax { font-size: 10px; color: #aaa; margin-left: 4px; }
.layout-6 .main-price-tax { font-size: 8px; }
.layout-9 .main-price-tax { display: none; }

.old-price {
  text-align: center;
  font-size: 12px;
  color: #aaa;
  margin-top: -6px;
  margin-bottom: 6px;
}
.old-price s { color: #bbb; }
.old-price .direction { font-weight: 700; margin-left: 6px; }
.old-price .direction.down { color: var(--blue); }
.old-price .direction.up { color: var(--red); }
.layout-4 .old-price { font-size: 9px; margin-top: -4px; }
.layout-6 .old-price { font-size: 8px; margin-top: -3px; margin-bottom: 3px; }
.layout-9 .old-price { font-size: 6px; margin-top: -2px; margin-bottom: 1px; }

/* サブ価格セクション */
.sub-prices { display: grid; gap: 0; }
.sub-row {
  display: flex;
  justify-content: space-between;
  align-items: center;
  padding: 5px 0;
  border-bottom: 1px dotted #e8e8e8;
  font-size: 13px;
}
.layout-4 .sub-row { padding: 3px 0; font-size: 10px; }
.layout-6 .sub-row { padding: 2px 0; font-size: 9px; }
.layout-9 .sub-row { padding: 1px 0; font-size: 7px; }
.sub-row:last-child { border-bottom: none; }
.sub-label { color: #777; font-weight: 500; }
.sub-value { font-weight: 700; color: var(--deep-brown); }
.sub-value.first { color: var(--blue); }
.sub-value.repeat { color: var(--purple); }
.sub-value.campaign { color: var(--red); }
.sub-value.monitor { color: var(--cyan); }

/* 回数券ボックス */
.ticket-box, .pp-box {
  margin-top: 8px;
  border: 1px solid #e5e5e5;
  border-radius: 8px;
  overflow: hidden;
}
.layout-4 .ticket-box, .layout-4 .pp-box { margin-top: 5px; border-radius: 5px; }
.layout-6 .ticket-box, .layout-6 .pp-box { margin-top: 4px; border-radius: 4px; }
.layout-9 .ticket-box, .layout-9 .pp-box { margin-top: 2px; border-radius: 3px; }
.ticket-header, .pp-header {
  background: #f8f8f8;
  padding: 4px 12px;
  font-size: 10px;
  font-weight: 700;
  color: #888;
}
.layout-4 .ticket-header, .layout-4 .pp-header { padding: 3px 8px; font-size: 8px; }
.layout-6 .ticket-header, .layout-6 .pp-header { padding: 2px 6px; font-size: 7px; }
.layout-9 .ticket-header, .layout-9 .pp-header { padding: 1px 5px; font-size: 6px; }
.ticket-row, .pp-row {
  display: flex;
  justify-content: space-between;
  padding: 4px 12px;
  font-size: 12px;
  border-top: 1px solid #f0f0f0;
}
.layout-4 .ticket-row, .layout-4 .pp-row { padding: 2px 8px; font-size: 9px; }
.layout-6 .ticket-row, .layout-6 .pp-row { padding: 2px 6px; font-size: 8px; }
.layout-9 .ticket-row, .layout-9 .pp-row { padding: 1px 5px; font-size: 6px; }
.ticket-row .tv, .pp-row .pv { font-weight: 700; }

.pp-box { border-color: var(--sage-light); }
.pp-header { background: var(--sage-light); color: var(--sage); }

/* extras（1行コンパクト表示） */
.extras {
  margin-top: 6px;
  padding: 5px 8px;
  background: #f9f9f9;
  border-radius: 6px;
  font-size: 10px;
  color: #777;
  line-height: 1.6;
}
.layout-6 .extras { font-size: 8px; margin-top: 4px; padding: 3px 6px; }
.layout-9 .extras { font-size: 6px; margin-top: 2px; padding: 2px 4px; }

/* バリアントテーブル */
.var-table {
  width: 100%;
  border-collapse: collapse;
  margin-top: 6px;
  font-size: 11px;
  table-layout: auto;
}
.layout-4 .var-table { font-size: 9px; margin-top: 4px; }
.layout-6 .var-table { font-size: 7.5px; margin-top: 3px; }
.layout-9 .var-table { font-size: 6px; margin-top: 2px; }

.var-table thead th {
  background: #f5f5f5;
  padding: 3px 4px;
  font-weight: 600;
  font-size: 0.9em;
  color: #999;
  border-bottom: 1.5px solid #ddd;
  text-align: left;
  white-space: nowrap;
}
.var-table thead th.vt-price { text-align: right; padding: 3px 2px; }

.var-table tbody td {
  padding: 3px 4px;
  border-bottom: 1px solid #f0f0f0;
  line-height: 1.3;
}
.layout-6 .var-table th, .layout-6 .var-table td { padding: 2px 3px; }
.layout-9 .var-table th, .layout-9 .var-table td { padding: 1px 2px; }

.var-table .vt-area {
  color: var(--deep-brown);
  font-weight: 500;
  max-width: 0;
  width: 40%;
  overflow: hidden;
  text-overflow: ellipsis;
}
.var-table .vt-price {
  text-align: right;
  font-weight: 600;
  white-space: nowrap;
  padding-left: 2px;
  padding-right: 2px;
  font-variant-numeric: tabular-nums;
  color: #444;
}
.var-table .vt-old { color: #aaa; font-weight: 400; }
.var-table .vt-new { font-weight: 800; color: var(--deep-brown); }
.var-table .price-down { color: var(--blue); font-weight: 800; }
.var-table .price-up { color: var(--red); font-weight: 800; }
.var-table thead th.vt-old-h { color: #bbb; }
.var-table thead th.vt-new-h { color: var(--deep-brown); }

/* テーブル内 割引列の色 */
.var-table .vt-first { color: var(--blue); }
.var-table .vt-repeat { color: var(--purple); }
.var-table .vt-campaign { color: var(--red); font-weight: 700; }
.var-table .vt-monitor { color: var(--cyan); }
.var-table .vt-ticket { color: #d97706; }
.var-table .vt-bp { color: #16a34a; }
.var-table .vt-ep { color: #0d9488; }

/* 割引色（単一バリアント） */
.ticket-row .tv { color: #d97706; }
.pp-row .pv-bp { color: #16a34a; }
.pp-row .pv-ep { color: #0d9488; }

/* 割引率バッジ（値の下に表示） */
.disc { display: block; font-size: 0.7em; color: #bbb; font-weight: 400; letter-spacing: -0.3px; line-height: 1.1; }
.layout-6 .disc { font-size: 0.65em; }
.layout-9 .disc { font-size: 0.6em; }

/* ========================================
   印刷用CSS — A4横に確実にフィット
   ======================================== */
@page {
  size: A4 portrait;
  margin: 6mm;
}

@media print {
  /* UI系を完全に消す */
  .no-print {
    display: none !important;
    visibility: hidden !important;
    height: 0 !important;
    overflow: hidden !important;
    position: absolute !important;
    width: 0 !important;
  }
  .preview-hint { display: none !important; }

  body {
    background: #fff !important;
    padding: 0 !important;
    margin: 0 !important;
  }
  .preview-area {
    padding: 0 !important;
    margin: 0 !important;
  }

  /* カードグリッド — 印刷用 */
  .card-grid {
    gap: 4mm;
    align-items: start;
  }
  .card-grid.layout-1 { max-width: none; grid-template-columns: 1fr; }
  .card-grid.layout-2 { grid-template-columns: repeat(2, 1fr); }
  .card-grid.layout-4 { grid-template-columns: repeat(2, 1fr); }
  .card-grid.layout-6 { grid-template-columns: repeat(3, 1fr); gap: 3mm; }
  .card-grid.layout-9 { grid-template-columns: repeat(3, 1fr); gap: 2mm; }

  .pop-card {
    border-width: 2px;
    break-inside: avoid;
    page-break-inside: avoid;
  }
}
</style>
</head>
<body>

<!-- コントロールパネル -->
<div class="no-print">
<div class="control-panel">
  <div class="control-row">
    <h1><span>BEAUTE</span> 施術POPジェネレーター</h1>
    <input type="text" class="search-box" id="searchBox" placeholder="🔍 施術名・部位で検索...">
  </div>
  <div class="cat-filter" id="catFilter"></div>
  <div class="action-row">
    <button class="layout-btn" data-layout="1" onclick="setLayout(1)">▢ 1枚</button>
    <button class="layout-btn" data-layout="2" onclick="setLayout(2)">◧ 2枚</button>
    <button class="layout-btn active" data-layout="4" onclick="setLayout(4)">⊞ 4枚</button>
    <button class="layout-btn" data-layout="6" onclick="setLayout(6)">▦ 6枚</button>
    <button class="layout-btn" data-layout="9" onclick="setLayout(9)">▩ 9枚</button>
    <span class="selected-count" id="selectedCount"></span>
    <button class="btn-clear" id="btnClear" onclick="clearSelection()" style="display:none">✕ 選択解除</button>
    <button class="btn-print" onclick="window.print()">🖨 印刷</button>
  </div>
</div>

<!-- 施術リスト -->
<div class="list-panel" id="listPanel">
  <div class="list-grid" id="listGrid"></div>
</div>
</div>

<!-- カードプレビュー -->
<div class="preview-area" id="previewArea">
  <div class="preview-hint no-print" id="previewHint">↑ 施術を選択するとPOPカードが表示されます</div>
  <div class="card-grid layout-4" id="cardGrid"></div>
</div>

<script>
// ─── データ埋め込み ───
const TREATMENTS = __TREATMENTS_JSON__;

// ─── 状態 ───
let selectedIds = new Set();
let currentLayout = 4;
let currentCategory = null;
let searchQuery = '';

// ─── 初期化 ───
function init() {
  buildCategoryFilter();
  buildList();
  updatePreview();
}

function buildCategoryFilter() {
  const cats = [...new Set(TREATMENTS.map(t => t.sheet))];
  const el = document.getElementById('catFilter');
  el.innerHTML = '<button class="cat-btn active" onclick="filterCategory(null, this)">すべて</button>' +
    cats.map(c => `<button class="cat-btn" onclick="filterCategory('${c}', this)">${c}</button>`).join('');
}

function filterCategory(cat, btn) {
  currentCategory = cat;
  document.querySelectorAll('.cat-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  buildList();
}

document.getElementById('searchBox').addEventListener('input', function() {
  searchQuery = this.value.trim().toLowerCase();
  buildList();
});

function getFilteredTreatments() {
  return TREATMENTS.filter(t => {
    if (currentCategory && t.sheet !== currentCategory) return false;
    if (searchQuery) {
      const haystack = (t.name + ' ' + t.area + ' ' + t.category + ' ' + t.unit).toLowerCase();
      return haystack.includes(searchQuery);
    }
    return true;
  });
}

function buildList() {
  const filtered = getFilteredTreatments();
  const el = document.getElementById('listGrid');
  el.innerHTML = filtered.map(t => {
    const sel = selectedIds.has(t.id) ? 'selected' : '';
    const v0 = t.variants[0];
    const price = v0.prices.price ? '¥' + v0.prices.price.toLocaleString() : '';
    const varCount = t.variants.length;
    const areaText = varCount > 1 ? `(${varCount}件)` : (v0.area || '');
    return `<div class="list-item ${sel}" onclick="toggleSelect(${t.id})" data-id="${t.id}">
      <div class="li-check">${sel ? '✓' : ''}</div>
      <span class="li-name">${t.name}</span>
      <span class="li-area">${areaText}</span>
      <span class="li-price">${price}${varCount > 1 ? '〜' : ''}</span>
    </div>`;
  }).join('');
}

function toggleSelect(id) {
  if (selectedIds.has(id)) {
    selectedIds.delete(id);
  } else {
    selectedIds.add(id);
  }
  buildList();
  updatePreview();
}

function clearSelection() {
  selectedIds.clear();
  buildList();
  updatePreview();
}

function setLayout(n) {
  currentLayout = n;
  document.querySelectorAll('.layout-btn').forEach(b => b.classList.remove('active'));
  document.querySelector(`.layout-btn[data-layout="${n}"]`).classList.add('active');
  updatePreview();
}

function fmt(v) {
  return '¥' + v.toLocaleString();
}
function fmtShort(v) {
  return '¥' + v.toLocaleString();
}
function discPct(base, val) {
  if (!base || !val || val >= base) return '';
  const pct = Math.round((1 - val / base) * 100);
  return pct > 0 ? '<span class="disc">' + pct + '%OFF</span>' : '';
}
function ticketUnit(key, val) {
  const n = {ticket_3:3, ticket_5:5, ticket_8:8}[key];
  if (!n || !val) return '';
  return '<span class="disc">1回あたり ' + fmt(Math.round(val / n)) + '</span>';
}

function buildCard(t) {
  const vars = t.variants;
  const isSingle = vars.length === 1;
  let html = `<div class="pop-card">`;

  // ヘッダー
  html += `<div class="card-header">
    <div class="card-clinic">BEAUTE SKIN CLINIC</div>
    <div class="card-category">${t.category}</div>
  </div>`;

  // 施術名
  html += `<div class="card-name">${t.name}</div>`;

  if (isSingle) {
    // ── 単一バリアント: 従来の大きい価格表示 ──
    const v = vars[0];
    const p = v.prices;
    if (v.area) html += `<div class="card-area">${v.area}</div>`;
    if (v.unit) html += `<div class="card-unit">${v.unit}</div>`;

    if (p.old_price && p.price && p.old_price !== p.price) {
      const dir = p.price < p.old_price ? 'down' : 'up';
      const arrow = dir === 'down' ? '↓' : '↑';
      html += `<div class="old-price">旧 <s>${fmt(p.old_price)}</s><span class="direction ${dir}">${arrow}</span></div>`;
    }

    html += `<div class="main-price">
      <div class="main-price-label">通常価格（税込）</div>
      <div class="main-price-value"><span class="main-price-yen">¥</span>${p.price.toLocaleString()}<span class="main-price-tax">（税込）</span></div>
    </div>`;

    let subs = [];
    const isEsteSingle = t.kubun === 'エステ';
    if (isEsteSingle && p.first) subs.push(['初回価格', fmt(p.first), 'first', p.first]);
    if (p.repeat) subs.push(['リピート', fmt(p.repeat), 'repeat', p.repeat]);
    if (p.campaign) subs.push(['CP', fmt(p.campaign), 'campaign', p.campaign]);
    if (subs.length) {
      html += '<div class="sub-prices">';
      subs.forEach(([l,v,c,n]) => html += `<div class="sub-row"><span class="sub-label">${l}</span><span class="sub-value ${c}">${v}${discPct(p.price, n)}</span></div>`);
      html += '</div>';
    }

    // 回数券
    let tickets = [];
    if (p.ticket_3) tickets.push(['3回券', fmt(p.ticket_3), p.ticket_3, 3]);
    if (p.ticket_5) tickets.push(['5回券', fmt(p.ticket_5), p.ticket_5, 5]);
    if (p.ticket_8) tickets.push(['8回券', fmt(p.ticket_8), p.ticket_8, 8]);
    if (tickets.length) {
      html += '<div class="ticket-box"><div class="ticket-header">回数券</div>';
      tickets.forEach(([l,v,n,times]) => {
        const unit = fmt(Math.round(n / times));
        html += `<div class="ticket-row"><span>${l}</span><span class="tv">${v}${discPct(p.price, n)}<span class="disc">1回あたり ${unit}</span></span></div>`;
      });
      html += '</div>';
    }

    // BP・EP（EPはエステのみ）
    const isEste = t.kubun === 'エステ';
    let pps = [];
    if (p.bp10) pps.push(['BP10万', fmt(p.bp10), 'bp', p.bp10]);
    if (p.bp30) pps.push(['BP30万', fmt(p.bp30), 'bp', p.bp30]);
    if (p.bp50) pps.push(['BP50万', fmt(p.bp50), 'bp', p.bp50]);
    if (isEste && p.ep5) pps.push(['EP5万', fmt(p.ep5), 'ep', p.ep5]);
    if (isEste && p.ep10) pps.push(['EP10万', fmt(p.ep10), 'ep', p.ep10]);
    if (isEste && p.ep20) pps.push(['EP20万', fmt(p.ep20), 'ep', p.ep20]);
    if (pps.length) {
      html += '<div class="pp-box"><div class="pp-header">プリペイド</div>';
      pps.forEach(([l,v,tp,n]) => html += `<div class="pp-row"><span>${l}</span><span class="pv pv-${tp}">${v}${discPct(p.price, n)}</span></div>`);
      html += '</div>';
    }

  } else {
    // ── 複数バリアント: テーブル形式 ──
    // 表示する価格列（データがある列のみ自動表示）
    const priceCols = [
      {key:'old_price', label:'旧通常'},
      {key:'old_first', label:'旧初回'},
      {key:'price', label:'通常'},
      {key:'first', label:'初回'},
      {key:'repeat', label:'リピート'},
      {key:'campaign', label:'CP'},
      {key:'moni_full', label:'モニ全'},
      {key:'moni_mask', label:'モニ隠'},
      {key:'moni_eye', label:'モニ目'},
      {key:'ticket_3', label:'3回券'},
      {key:'ticket_5', label:'5回券'},
      {key:'ticket_8', label:'8回券'},
      {key:'bp10', label:'BP10'},
      {key:'bp30', label:'BP30'},
      {key:'bp50', label:'BP50'},
      {key:'ep5', label:'EP5'},
      {key:'ep10', label:'EP10'},
      {key:'ep20', label:'EP20'},
    ];
    // どの列にデータがあるか（EP・初回はエステのみ）
    const isEste = t.kubun === 'エステ';
    const activeCols = priceCols.filter(col => {
      if (!isEste && col.key.startsWith('ep')) return false;
      if (!isEste && col.key === 'first') return false;
      return vars.some(v => v.prices[col.key]);
    });

    html += '<table class="var-table"><thead><tr>';
    html += '<th class="vt-area"></th>';
    activeCols.forEach(col => {
      let hcls = 'vt-price';
      if (col.key === 'old_price' || col.key === 'old_first') hcls += ' vt-old-h';
      if (col.key === 'price') hcls += ' vt-new-h';
      html += `<th class="${hcls}">${col.label}</th>`;
    });
    html += '</tr></thead><tbody>';

    vars.forEach(v => {
      const p = v.prices;
      const label = v.area || v.unit || '-';
      const changed = p.old_price && p.price && p.old_price !== p.price;
      html += '<tr>';
      html += `<td class="vt-area">${label}</td>`;
      activeCols.forEach(col => {
        const val = p[col.key];
        let cls = 'vt-price';
        if (col.key === 'old_price' || col.key === 'old_first') cls += ' vt-old';
        else if (col.key === 'price') {
          cls += ' vt-new';
          if (changed) cls += p.price < p.old_price ? ' price-down' : ' price-up';
        }
        else if (col.key === 'first') cls += ' vt-first';
        else if (col.key === 'repeat') cls += ' vt-repeat';
        else if (col.key === 'campaign') cls += ' vt-campaign';
        else if (col.key.startsWith('moni')) cls += ' vt-monitor';
        else if (col.key.startsWith('ticket')) cls += ' vt-ticket';
        else if (col.key.startsWith('bp')) cls += ' vt-bp';
        else if (col.key.startsWith('ep')) cls += ' vt-ep';
        let cell = val ? fmtShort(val) : '';
        if (val && p.price && val < p.price && col.key !== 'old_price' && col.key !== 'old_first' && col.key !== 'price') {
          cell += discPct(p.price, val);
        }
        if (val && col.key.startsWith('ticket')) {
          cell += ticketUnit(col.key, val);
        }
        html += `<td class="${cls}">${cell}</td>`;
      });
      html += '</tr>';
    });

    html += '</tbody></table>';
  }

  html += '</div>';
  return html;
}

function updatePreview() {
  const count = selectedIds.size;
  document.getElementById('selectedCount').textContent = count ? `${count}件選択中` : '';
  document.getElementById('btnClear').style.display = count ? '' : 'none';
  document.getElementById('previewHint').style.display = count ? 'none' : '';

  const selected = TREATMENTS.filter(t => selectedIds.has(t.id));
  const container = document.getElementById('cardGrid');
  container.className = 'card-grid layout-' + currentLayout;

  if (!selected.length) {
    container.innerHTML = '';
    return;
  }

  container.innerHTML = selected.map(t => buildCard(t)).join('');
}

init();
</script>
</body>
</html>'''

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 build_pop.py <input.xlsx> [output.html]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(input_path) or '.', '施術POP.html'
    )

    print(f"📖 Reading: {input_path}")
    treatments = read_excel(input_path)
    print(f"   → {len(treatments)} treatments loaded")

    # JSON埋め込み
    json_data = json.dumps(treatments, ensure_ascii=False, separators=(',', ':'))
    html = HTML_TEMPLATE.replace('__TREATMENTS_JSON__', json_data)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"✅ Generated: {output_path}")
    print(f"   ブラウザで開いて施術を選択 → 印刷してください")

if __name__ == '__main__':
    main()
